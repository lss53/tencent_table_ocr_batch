#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

依赖腾讯云的「表格识别（V3）」模块，实现特定图片转表格功能

- Author: DS
- Created Time: 2025/06/20 18:52
- Copyright: Copyright © 2025 DS. All rights reserved.
"""

import json
import base64
import os
import time
import argparse
import concurrent.futures
import logging
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.ocr.v20181119 import ocr_client, models
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# 支持的图片格式
SUPPORTED_EXT = ['.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff']
MAX_IMAGE_SIZE = 3 * 1024 * 1024  # 3MB

# 配置日志
def setup_logger(log_dir=None):
    logger = logging.getLogger('TableOCR')
    
    # 清除现有的处理器（避免重复添加）
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    logger.setLevel(logging.INFO)
    
    # 创建统一的日志格式
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # 创建控制台处理器 (始终启用)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    
    # 如果指定了日志文件，则添加文件处理器
    if log_dir:
        os.makedirs(log_dir, exist_ok=True)
        
        # 生成带时间戳的日志文件名
        timestamp = time.strftime("%Y%m%d")
        log_file = os.path.join(log_dir, f"table_ocr_{timestamp}.log")
        
        # 检查日志文件是否存在且非空
        log_exists = os.path.exists(log_file) and os.path.getsize(log_file) > 0
        
        # 创建文件处理器(追加模式)
        file_handler = logging.FileHandler(log_file, encoding='utf-8', mode='a')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
       # 如果日志文件已存在且非空，添加分隔符
        if log_exists:
            # 添加2行分隔符
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write('\n' * 2)
    
    return logger

def log_divider(title=None):
    """输出分隔线"""
    logger = logging.getLogger('TableOCR')
    if title:
        logger.info(f"{'=' * 60}")
        logger.info(title.center(50))
        logger.info(f"{'=' * 60}")
    else:
        logger.info(f"{'=' * 60}")

def check_image_size(image_path):
    """检查图片大小是否超过限制"""
    logger = logging.getLogger('TableOCR')
    try:
        file_size = os.path.getsize(image_path)
        if file_size > MAX_IMAGE_SIZE:
            logger.warning(f"图片 {os.path.basename(image_path)} 超过3MB限制 ({file_size/1024/1024:.2f}MB)，已跳过")
            return False
        return True
    except Exception as e:
        logger.error(f"检查图片大小时出错: {str(e)}")
        return False

def process_image(image_path, client):
    """处理单张图片识别并返回结果（带智能重试机制和详细错误分类）"""
    logger = logging.getLogger('TableOCR')
    try:
        # 检查图片大小
        if not check_image_size(image_path):
            return None, "图片超过3MB限制"
            
        # 读取图片并转换为Base64
        with open(image_path, "rb") as image_file:
            image_base64 = base64.b64encode(image_file.read()).decode('utf-8')
        
        # 构建请求参数
        req = models.RecognizeTableAccurateOCRRequest()
        params = {"ImageBase64": image_base64}
        req.from_json_string(json.dumps(params))
        
        # 错误分类配置
        non_retryable_errors = {
            # 认证错误
            "AuthFailure.SecretIdNotFound": "密钥ID不存在",
            "AuthFailure.SignatureFailure": "签名验证失败",
            "AuthFailure.TokenFailure": "临时令牌无效",
            
            # 账户问题
            "FailedOperation.ArrearsError": "账户欠费",
            "FailedOperation.UnOpenError": "服务未开通",
            "LimitExceeded": "超过配额限制",
            
            # 图片问题
            "FailedOperation.OcrFailed.InvalidImage": "无效图片格式",
            "FailedOperation.OcrFailed.TooLarge": "图片超过3MB限制",
            "FailedOperation.OcrFailed.UnsupportedFormat": "不支持的图片格式",
            "FailedOperation.OcrFailed.LowQuality": "图片质量过低",
            "FailedOperation.OcrFailed.NoText": "未检测到文本",
            "FailedOperation.OcrFailed.NoTable": "未检测到表格",
            "FailedOperation.OcrFailed.ComplexTable": "表格结构过于复杂",
            "FailedOperation.ImageSizeTooSmall": "图片尺寸过小",
            "FailedOperation.ImageNoText": "图片中无文本",
            
            # 参数错误
            "InvalidParameter": "参数错误",
            "InvalidParameterValue": "参数值无效",
            "MissingParameter": "缺少必要参数"
        }
        
        # 错误诊断建议
        error_guidance = {
            "FailedOperation.OcrFailed.LowQuality": "建议：提高图片分辨率，增强对比度，减少反光",
            "FailedOperation.OcrFailed.ComplexTable": "建议：简化表格结构，避免嵌套表格，减少合并单元格",
            "FailedOperation.OcrFailed.NoText": "建议：确认图片包含文字，检查文字方向",
            "FailedOperation.OcrFailed.NoTable": "建议：确保图片包含清晰的表格边框",
            "FailedOperation.ImageDecodeFailed": "建议：重新保存图片为标准JPEG/PNG格式"
        }
        
        # 重试机制配置
        max_retries = 3
        retry_delay = 2  # 初始延迟2秒
        request_id = None
        
        for attempt in range(max_retries):
            try:
                # 发送请求
                resp = client.RecognizeTableAccurateOCR(req)
                return json.loads(resp.to_json_string()), None
                
            except TencentCloudSDKException as err:
                error_code = err.get_code()
                error_msg = err.get_message()
                request_id = err.get_request_id()
                
                # 记录详细的错误信息
                logger.debug(f"OCR请求错误: [代码] {error_code}, [消息] {error_msg}, [请求ID] {request_id}")
                
                # 1. 处理不可重试的错误
                if error_code in non_retryable_errors:
                    friendly_msg = non_retryable_errors[error_code]
                    suggestion = error_guidance.get(error_code, "建议：检查图片内容并重试")
                    return None, f"{friendly_msg} | {suggestion} [错误码: {error_code}]"
                
                # 2. 处理认证错误（特殊处理）
                if error_code.startswith("AuthFailure"):
                    raise RuntimeError(f"腾讯云认证错误: {error_msg} [错误码: {error_code}]") from err
                
                # 3. 可重试错误处理
                if attempt < max_retries - 1:
                    # 错误类型分类
                    if "LimitExceeded" in error_code:
                        error_type = "限频错误"
                    elif "Internal" in error_code or "UnKnow" in error_code:
                        error_type = "服务端错误"
                    else:
                        error_type = "可恢复错误"
                    
                    logger.warning(
                        f"图片 {os.path.basename(image_path)} 识别失败 ({error_type}, 尝试 {attempt+1}/{max_retries}): "
                        f"[{error_code}] {error_msg} - {retry_delay}秒后重试"
                    )
                    time.sleep(retry_delay)
                    retry_delay *= 2  # 指数退避
                    continue
                
                # 达到最大重试次数仍失败
                return None, f"重试失败: {error_msg} [错误码: {error_code}]"
                
            except Exception as e:
                # 网络错误处理
                if attempt < max_retries - 1:
                    logger.warning(
                        f"图片 {os.path.basename(image_path)} 网络错误 (尝试 {attempt+1}/{max_retries}): "
                        f"{str(e)} - {retry_delay}秒后重试"
                    )
                    time.sleep(retry_delay)
                    retry_delay *= 2
                    continue
                
                # 达到最大重试次数仍失败
                return None, f"网络错误: {str(e)}"
        
        return None, "达到最大重试次数"
        
    except Exception as e:
        # 捕获处理过程中的其他异常
        logger.error(f"处理图片 {os.path.basename(image_path)} 时发生意外错误: {str(e)}")
        return None, f"处理错误: {str(e)}"

def create_worksheet(wb, base_name, result):
    """为识别结果创建工作表"""
    logger = logging.getLogger('TableOCR')
    try:
        # 检查是否有有效的表格识别结果
        if 'TableDetections' not in result or len(result['TableDetections']) < 2:
            return False, "未识别到有效的表格数据"
        
        # 创建新的工作表
        ws = wb.create_sheet(title=base_name)
        
        # 提取主表格（第二个表格，索引1）
        main_table = result['TableDetections'][1]
        cells = main_table['Cells']

        # 设置列宽
        column_widths = [8, 25, 8, 40, 10, 10, 10, 10, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 预计算最大行列
        max_row = max(cell['RowBr'] for cell in cells)
        max_col = max(cell['ColBr'] for cell in cells)
        
        # 初始化单元格字典
        cell_dict = {}
        
        # 处理合并单元格和填充数据
        merged_cells = []
        
        for cell in cells:
            row_start = cell['RowTl'] + 1  # Excel行号从1开始
            row_end = cell['RowBr']        # 开区间，直接作为结束行
            col_start = cell['ColTl'] + 1  # Excel列号从1开始
            col_end = cell['ColBr']        # 开区间，直接作为结束列
            
            # 处理文本中的换行符
            text = cell['Text'].replace('\n', '')
            
            # 只填充左上角单元格
            cell_dict[(row_start, col_start)] = text
            
            # 记录需要合并的单元格
            if row_end > row_start or col_end > col_start:
                merged_cells.append((row_start, row_end, col_start, col_end))

        # 批量填充单元格
        for (row, col), value in cell_dict.items():
            ws.cell(row=row, column=col, value=value)
        
        # 应用合并单元格
        for merge in merged_cells:
            row_start, row_end, col_start, col_end = merge
            ws.merge_cells(
                start_row=row_start, end_row=row_end,
                start_column=col_start, end_column=col_end
            )

        # 设置居中和自动换行 - 只设置有内容的单元格
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if (row, col) in cell_dict:
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(
                        horizontal='center', 
                        vertical='center',
                        wrap_text=True
                    )
        return True, None
    except Exception as e:
        return False, f"创建工作表出错: {str(e)}"

def save_workbook(wb, output_path, batch_num=None):
    """保存工作簿到指定路径"""
    logger = logging.getLogger('TableOCR')
    try:
        if batch_num is not None:
            # 添加批次号到临时文件名
            base_name, ext = os.path.splitext(output_path)
            temp_path = f"{base_name}_batch{batch_num}{ext}"
            wb.save(temp_path)
            logger.info(f"临时结果已保存: {os.path.basename(temp_path)}")
        else:
            wb.save(output_path)
            logger.info(f"最终结果已保存: {os.path.basename(output_path)}")
        return True
    except Exception as e:
        logger.error(f"保存工作簿时出错: {str(e)}")
        return False

def main():
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='腾讯云表格OCR批量处理工具')
    parser.add_argument('--image_dir', required=True, help='图片文件夹路径')
    parser.add_argument('--output_dir', required=True, help='结果保存路径')
    parser.add_argument('--secret_id', required=True, help='腾讯云API密钥ID')
    parser.add_argument('--secret_key', required=True, help='腾讯云API密钥KEY')
    parser.add_argument('--max_workers', type=int, default=2, help='最大并发线程数 (默认: 2)')
    parser.add_argument('--batch_size', type=int, default=10, help='每批处理完成后保存的图片数量 (默认: 10)')
    parser.add_argument('--region', default='ap-chongqing', help='服务地域 (默认: ap-chongqing)')
    parser.add_argument('--log_dir', help='将日志文件记录到指定路径')
    args = parser.parse_args()
    
    # 初始化日志记录器
    setup_logger(args.log_dir)
    logger = logging.getLogger('TableOCR')
    
    # 创建输出目录
    os.makedirs(args.output_dir, exist_ok=True)
    
    try:
        log_divider(f"腾讯云表格OCR批量处理工具 - {time.strftime('%Y-%m-%d %H:%M:%S')}")
        # 初始化认证和客户端
        cred = credential.Credential(args.secret_id, args.secret_key)
        httpProfile = HttpProfile()
        httpProfile.endpoint = "ocr.tencentcloudapi.com"
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        client = ocr_client.OcrClient(cred, args.region, clientProfile)
        
        # 获取所有支持的图片文件
        image_files = [
            os.path.join(args.image_dir, f) for f in os.listdir(args.image_dir) 
            if os.path.splitext(f)[1].lower() in SUPPORTED_EXT
        ]
        image_files.sort()  # 确保文件按顺序处理
        
        if not image_files:
            logger.warning(f"在 {args.image_dir} 目录中未找到支持的图片文件")
            return
        
        logger.info(f"使用最大并发数: {args.max_workers}")
        logger.info(f"服务地域: {args.region}")
        logger.info(f"图片目录: {os.path.abspath(args.image_dir)}")
        logger.info(f"结果目录: {os.path.abspath(args.output_dir)}")
        if args.log_dir:
            logger.info(f"日志目录: {os.path.abspath(args.log_dir)}")
        logger.info(f"找到 {len(image_files)} 张图片待处理...")
        logger.info(f"每 {args.batch_size} 张图片保存一次临时结果")
        log_divider()
        
        # 获取图片目录的basename作为输出文件名
        dir_basename = os.path.basename(os.path.normpath(args.image_dir))
        if not dir_basename or dir_basename == '.':
            dir_basename = "table_ocr_result"
        
        # 初始化工作簿（但不立即创建）
        wb = None
        output_file = f"{dir_basename}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(args.output_dir, output_file)
        
        # 记录开始时间
        start_time = time.time()
        processed_count = 0
        success_count = 0
        batch_counter = 0
        failed_files = []  # 存储失败文件信息
        
        # 使用线程池并行处理图片
        with concurrent.futures.ThreadPoolExecutor(max_workers=args.max_workers) as executor:
            # 提交所有任务
            future_to_image = {executor.submit(process_image, img_path, client): img_path for img_path in image_files}
            
            # 处理完成的任务
            for future in concurrent.futures.as_completed(future_to_image):
                img_path = future_to_image[future]
                processed_count += 1
                img_name = os.path.basename(img_path)
                
                try:
                    result, error_msg = future.result()
                    if result:
                        # 提取图片基本名称（不含扩展名）作为工作表名
                        base_name = os.path.splitext(img_name)[0]
                        
                        # 第一次成功识别时创建工作簿
                        if wb is None:
                            wb = openpyxl.Workbook()
                            # 删除默认创建的工作表
                            if 'Sheet' in wb.sheetnames:
                                del wb['Sheet']
                            logger.info("已创建新的工作簿")
                        
                        # 创建工作表
                        success, create_error = create_worksheet(wb, base_name, result)
                        if success:
                            success_count += 1
                            batch_counter += 1
                            logger.info(f"已处理 ({processed_count}/{len(image_files)}): {img_name} [成功]")
                            
                            # 分批保存
                            if batch_counter >= args.batch_size and wb is not None:
                                save_workbook(wb, output_path, success_count // args.batch_size)
                                batch_counter = 0
                        else:
                            failed_files.append((img_name, create_error))
                            logger.warning(f"已处理 ({processed_count}/{len(image_files)}): {img_name} [识别失败] - {create_error}")
                    else:
                        failed_files.append((img_name, error_msg))
                        logger.warning(f"已处理 ({processed_count}/{len(image_files)}): {img_name} [失败] - {error_msg}")
                except RuntimeError as e:
                    # 处理认证错误
                    logger.critical(f"致命错误: {str(e)}")
                    logger.critical("程序终止")
                    return
                except Exception as e:
                    failed_files.append((img_name, str(e)))
                    logger.error(f"处理图片 {img_name} 时出错: {str(e)}")
        
        # 最终保存工作簿（如果有工作簿）
        if wb is not None:
            save_workbook(wb, output_path)
        
        # 计算处理时间
        total_time = time.time() - start_time
        
        # 创建美观的总结输出
        log_divider("处理结果统计")
        logger.info(f"总图片数: {len(image_files)}")
        logger.info(f"成功识别数: {success_count}")
        logger.info(f"失败数: {len(failed_files)}")
        logger.info(f"总耗时: {total_time:.2f}秒")
        logger.info(f"平均每张图片: {total_time/len(image_files):.2f}秒")
        
        # 输出失败文件列表
        # if failed_files:
        #     logger.info("失败文件列表:")
        #     for file_name, reason in failed_files:
        #         logger.info(f" - {file_name}: {reason}")
        if failed_files:
            logger.info("失败文件列表:")
            # 生成失败文件报告
            fail_report_path = os.path.join(args.output_dir, f"{dir_basename}_failures_{time.strftime('%Y%m%d_%H%M%S')}.txt")
            with open(fail_report_path, 'w', encoding='utf-8') as f:
                f.write("失败文件列表:\n")
                f.write("="*50 + "\n")
                for i, (file_name, reason) in enumerate(failed_files, 1):
                    f.write(f"{i}. {file_name}: {reason}\n")
            
            # logger.info(f" - 共 {len(failed_files)} 个失败文件")
            logger.info(f" - 失败详情已导出至: {os.path.basename(fail_report_path)}")
        log_divider()
        
        if success_count > 0:
            logger.info(f"Excel文件已生成: {os.path.abspath(output_path)}")
        else:
            logger.warning("没有成功识别的图片，未生成Excel文件")
        
    except TencentCloudSDKException as err:
        logger.error(f"腾讯云SDK错误: {err}")
    except Exception as e:
        logger.critical(f"程序异常: {str(e)}")

if __name__ == "__main__":
    main()